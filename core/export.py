"""Export builders — XLSX, htaccess CSV, JSON."""

from __future__ import annotations

import io
import json

import pandas as pd


def build_review_xlsx(
    results_df: pd.DataFrame,
    ai_df: pd.DataFrame | None = None,
    mode: str = "migration",
) -> bytes:
    """Build multi-sheet review XLSX.

    Sheets: Best Matches, High Confidence, Needs Review, No Match,
    per-method sheets, and AI Decisions (if ai_df provided).
    """
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_sheet(writer, results_df, "Best Matches")

        if "tier" in results_df.columns:
            high = results_df[results_df["tier"] == "high"]
            review = results_df[results_df["tier"] == "review"]
            no_match = results_df[results_df["tier"] == "no_match"]
        else:
            score_col = "combined_score" if "combined_score" in results_df.columns else "score"
            high = results_df[results_df.get(score_col, pd.Series(dtype=float)) >= 0.90] if score_col in results_df.columns else results_df.iloc[0:0]
            review = results_df[(results_df.get(score_col, pd.Series(dtype=float)) >= 0.70) & (results_df.get(score_col, pd.Series(dtype=float)) < 0.90)] if score_col in results_df.columns else results_df.iloc[0:0]
            no_match = results_df[results_df.get(score_col, pd.Series(dtype=float)) < 0.70] if score_col in results_df.columns else results_df.iloc[0:0]

        _write_sheet(writer, high, "High Confidence")
        _write_sheet(writer, review, "Needs Review")
        _write_sheet(writer, no_match, "No Match")

        # Per-method sheets
        method_col = "methods_contributed" if "methods_contributed" in results_df.columns else None
        if method_col:
            for method in results_df[method_col].dropna().unique():
                subset = results_df[results_df[method_col].str.contains(str(method), na=False)]
                sheet_name = f"By Method: {str(method)[:22]}"
                _write_sheet(writer, subset, sheet_name)

        if ai_df is not None and not ai_df.empty:
            _write_sheet(writer, ai_df, "AI Decisions")

    return buf.getvalue()


def build_high_confidence_csv(results_df: pd.DataFrame) -> bytes:
    """Two-column CSV: source_url, destination_url for tier == 'high' rows."""
    if results_df.empty:
        return b"source_url,destination_url\n"

    if "tier" in results_df.columns:
        subset = results_df[results_df["tier"] == "high"]
    else:
        score_col = "combined_score" if "combined_score" in results_df.columns else "score"
        subset = results_df[results_df[score_col] >= 0.90] if score_col in results_df.columns else results_df.iloc[0:0]

    src_col = "legacy_url" if "legacy_url" in subset.columns else subset.columns[0]
    dst_col = "candidate_url" if "candidate_url" in subset.columns else subset.columns[1]

    out = subset[[src_col, dst_col]].rename(
        columns={src_col: "source_url", dst_col: "destination_url"}
    )
    return out.to_csv(index=False).encode("utf-8")


def build_json(
    results_df: pd.DataFrame,
    ai_df: pd.DataFrame | None = None,
) -> bytes:
    """Full JSON dump of results and AI decisions."""
    payload = {
        "results": results_df.to_dict(orient="records"),
        "ai_decisions": ai_df.to_dict(orient="records") if ai_df is not None else [],
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    """Write a DataFrame to a sheet with bold headers, frozen top row, auto-sized columns."""
    safe_name = sheet_name[:31]
    df.to_excel(writer, sheet_name=safe_name, index=False)

    ws = writer.sheets[safe_name]
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = ws["A2"]

    for col_idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
